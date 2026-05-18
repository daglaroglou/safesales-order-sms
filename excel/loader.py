from __future__ import annotations

from pathlib import Path

import openpyxl

from excel.acs import ACS_SHEET_NAME, parse_acs_file
from excel.box_express import BOX_EXPRESS_SHEET_NAME, parse_box_express_file
from excel.models import Shipment
from excel.utils import excel_dir, is_generated_export


def detect_parser(path: Path) -> str | None:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if ACS_SHEET_NAME in workbook.sheetnames:
            return "acs"
        if BOX_EXPRESS_SHEET_NAME in workbook.sheetnames:
            return "box_express"
    finally:
        workbook.close()

    return None


def parse_excel_file(
    path: Path | str,
    *,
    export_directory: Path | str | None = None,
) -> list[Shipment]:
    source = Path(path)
    parser = detect_parser(source)
    export_dir = Path(export_directory) if export_directory is not None else source.parent

    if parser == "acs":
        from excel.utils import acs_output_path

        return parse_acs_file(source, output_path=acs_output_path(export_dir))
    if parser == "box_express":
        from excel.utils import box_express_output_path

        return parse_box_express_file(source, output_path=box_express_output_path(export_dir))

    raise ValueError(f"Unsupported Excel file: {source.name}")


def parse_excel_folder(directory: Path | str | None = None) -> list[Shipment]:
    folder = Path(directory) if directory is not None else excel_dir()
    shipments: list[Shipment] = []

    for path in sorted(folder.glob("*.xlsx")):
        if is_generated_export(path):
            continue
        shipments.extend(parse_excel_file(path))

    return shipments
