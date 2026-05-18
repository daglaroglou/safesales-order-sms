from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import openpyxl


def excel_dir() -> Path:
    return Path(__file__).resolve().parent


def format_export_date(value: date | None = None) -> str:
    current = value or date.today()
    return f"{current.day}-{current.month}-{current.year}"


def acs_output_path(directory: Path | str, on_date: date | None = None) -> Path:
    return Path(directory) / f"{format_export_date(on_date)}.xlsx"


def box_express_output_path(directory: Path | str, on_date: date | None = None) -> Path:
    return Path(directory) / f"{format_export_date(on_date)}be.xlsx"


def is_generated_export(path: Path | str) -> bool:
    return bool(re.fullmatch(r"\d+-\d+-\d+(?:be)?\.xlsx", Path(path).name))


def write_pair_workbook(
    path: Path | str,
    sheet_name: str,
    pairs: Iterable[tuple[str, str]],
) -> Path:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    for voucher, phone in pairs:
        worksheet.append([voucher, phone])

    workbook.save(destination)
    return destination


def normalize_phone(value: Any) -> str | None:
    if value is None:
        return None

    digits = "".join(character for character in str(value).strip() if character.isdigit())
    if not digits:
        return None

    if len(digits) == 10 and digits.startswith("69"):
        return f"30{digits}"
    if len(digits) == 11 and digits.startswith("069"):
        return f"30{digits[1:]}"
    if digits.startswith("30"):
        return digits

    return digits


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def normalize_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def normalize_date(value: Any) -> date | datetime | None:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    return None


def iter_sheet_rows(path: Path, sheet_name: str | None = None) -> Iterable[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [normalize_text(header) for header in next(rows, ())]
        if not headers or all(header is None for header in headers):
            return

        for row in rows:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if header is None:
                    continue
                record[header] = row[index] if index < len(row) else None
            yield record
    finally:
        workbook.close()
